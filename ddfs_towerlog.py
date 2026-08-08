"""ddfs_towerlog.py - TOWERLOG / AODB BASE-COMPLETENESS RECONCILIATION
(23 July 2026, ACI review item 2, P1).

The OAG-based DDFS base is schedules only: cargo, GA, military, positioning
and ad-hoc charter never appear, so for an airport heavy in those the design
day understates the true peak. The Avia gold standard (Bologna) starts from
towerlogs, which carry every movement. This module ingests a towerlog day
(MZLZ AODB export format: Jess's "2025 Towerlog.xlsx", full-year ZAG actuals)
and reconciles it against the OAG schedule day side by side, with the
movement gap quantified BY CATEGORY, never netted silently.

Movement classes (from the log's own Categ. and Kind fields):
  scheduled_pax  Kind PA, Categ S      charter_pax  Kind PA, Categ C
  ga             Categ G (Kind PA)     cargo        Kind C
  positioning    Kind ELP              military     Kind MC
  other          anything else, kept and labelled

Sources: 2025 Towerlog.xlsx (Egnyte, MZLZ Full Forecast 01 Data and Analysis;
local working copy C:\\Avia\\2025 Towerlog.xlsx when placed). Columns per the
export header: Dir. | Rwy date | Rwy.t. | C/P | Registr. | Op.airl./usr |
Line | Sto.1 | Sto.1 IATA | ... | Aircr.typ. | Categ. | ... | MTOW | Pax cap.
| ... | Kind | ... Pay.

Run:  python3 ddfs_towerlog.py --selftest
      python3 ddfs_towerlog.py --extract "C:\\Avia\\2025 Towerlog.xlsx" 2025-05-26
      python3 ddfs_towerlog.py --reconcile ZAG 2025-05-26
Author: Avia Solutions.
"""
import os, csv, sys, glob, datetime as dt
from collections import defaultdict

FIX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ddfs_bridge_fixtures")

# ICAO aerodrome reference letter by towerlog aircraft type (wingspan-based;
# extends the OAG map with GA / military / bizjet types the log carries)
ICAO_TL = {"A320": "C", "A319": "C", "A321": "C", "A20N": "C", "A21N": "C",
           "B738": "C", "B38M": "C", "BCS3": "C", "A220": "C",
           "E170": "C", "E190": "C", "E195": "C", "E295": "C", "CRJX": "B",
           "DH8D": "C", "AT72": "C", "SF3": "B", "PC12": "A", "PA28": "A",
           "C25A": "A", "C525": "A", "C56X": "B", "E55P": "A", "GLF5": "C",
           "C130": "C", "AT45": "C", "B737": "C", "E75L": "C"}


def classify(categ, kind):
    kind = (kind or "").strip().upper()
    categ = (categ or "").strip().upper()
    if kind == "ELP":
        return "positioning"
    if kind == "MC":
        return "military"
    if kind == "C":
        return "cargo"
    if categ == "G":
        return "ga"
    if kind == "PA" and categ == "C":
        return "charter_pax"
    if kind == "PA" and categ == "S":
        return "scheduled_pax"
    return "other"


def _minute_of(v):
    """Runway time as minutes: 'H:MM' string or an Excel time cell
    (datetime with 1900 date-part)."""
    if isinstance(v, dt.datetime):
        return v.hour * 60 + v.minute
    s = str(v).strip()
    if ":" not in s:
        return None
    hh, mm = s.split(":")[:2]
    try:
        return int(hh) * 60 + int(mm)
    except ValueError:
        return None


def _parse_rows(rows, date):
    """rows in the export column order, with or without the leading 'Sheet'
    column (the workbook has none; the Egnyte text extraction carries one):
    Dir | Rwy date | Rwy.t. | C/P | Registr. | Op.airl | Line | Sto.1 |
    Sto.1 IATA | Sto.2 | Sto.2 IATA | Aircr.typ. | Categ. | Compl.code |
    MTOW | Pax cap. | ... | Kind(+29) | ... Pay.(+32). Dates arrive as
    datetime cells or M/D/YY strings; both handled, nothing inferred."""
    want = f"{date.month}/{date.day}/{str(date.year)[2:]}"
    out = []
    for r in rows:
        o = 0 if str(r[0]).strip() in ("A", "D") else (1 if len(r) > 1 and str(r[1]).strip() in ("A", "D") else None)
        if o is None or len(r) < o + 30:
            continue
        if str(r[o + 1]).strip() != want and not _same_date(r[o + 1], date):
            continue
        minute = _minute_of(r[o + 2])
        if minute is None:
            continue
        actype = str(r[o + 11]).strip()
        pax = None
        if len(r) > o + 32:
            try:
                pax = int(float(str(r[o + 32]).strip()))
            except (ValueError, TypeError):
                pax = None
        out.append(dict(
            ad=str(r[o]).strip(), minute=minute,
            registration=str(r[o + 4]).strip(), carrier=str(r[o + 5]).strip(),
            fno=str(r[o + 6]).strip(), od=str(r[o + 8]).strip(), actype=actype,
            icao=ICAO_TL.get(actype, "C"),
            categ=str(r[o + 12]).strip(), kind=str(r[o + 29]).strip(),
            cls=classify(r[o + 12], r[o + 29]), pax=pax))
    return sorted(out, key=lambda e: e["minute"])


def _same_date(v, date):
    if isinstance(v, dt.datetime):
        return v.date() == date
    if isinstance(v, dt.date):
        return v == date
    return False


def load_xlsx_day(path, date):
    """One calendar day from the towerlog workbook (openpyxl, first sheet)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ([("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True))
    ev = _parse_rows(rows, date)
    wb.close()
    return ev


def load_fixture_day(name):
    """A pinned towerlog day fixture (TSV in the export column subset:
    dir, time, registration, carrier, flight, od_iata, actype, categ, kind, pax)."""
    ev = []
    with open(os.path.join(FIX, name)) as f:
        for r in csv.DictReader((l for l in f if not l.startswith("#")), delimiter="\t"):
            hh, mm = r["time"].split(":")[:2]
            pax = r.get("pax") or ""
            ev.append(dict(
                ad=r["dir"], minute=int(hh) * 60 + int(mm),
                registration=r.get("registration", ""), carrier=r["carrier"],
                fno=r["flight"], od=r["od_iata"], actype=r["actype"],
                icao=ICAO_TL.get(r["actype"], "C"),
                categ=r["categ"], kind=r["kind"],
                cls=classify(r["categ"], r["kind"]),
                pax=int(pax) if pax.strip().isdigit() else None))
    return sorted(ev, key=lambda e: e["minute"])


def extract_day(xlsx_path, date, out_name=None):
    """Pin one towerlog day as a fixture TSV (flag rather than fill: every
    row carries its class; nothing is dropped)."""
    ev = load_xlsx_day(xlsx_path, date)
    out_name = out_name or f"zag_towerlog_{date.isoformat()}.tsv"
    p = os.path.join(FIX, out_name)
    with open(p, "w") as f:
        f.write(f"# Towerlog day {date.isoformat()}, extracted from {os.path.basename(xlsx_path)}\n")
        f.write("# Source: MZLZ AODB towerlog (Jess's 2025 Towerlog.xlsx); every movement carried, classes per the log's Categ./Kind fields\n")
        f.write("dir\ttime\tregistration\tcarrier\tflight\tod_iata\tactype\tcateg\tkind\tpax\n")
        for e in ev:
            f.write(f"{e['ad']}\t{e['minute']//60:02d}:{e['minute']%60:02d}\t{e['registration']}\t{e['carrier']}"
                    f"\t{e['fno']}\t{e['od']}\t{e['actype']}\t{e['categ']}\t{e['kind']}"
                    f"\t{'' if e['pax'] is None else e['pax']}\n")
    print(f"pinned {len(ev)} movements to {p}")
    return p


def completeness_statement(base, excludes):
    """The base-and-exclusions statement every output carries (item 2)."""
    return (f"BASE: {base}. EXCLUDES: {', '.join(excludes)}. "
            "Where a towerlog or AODB day is held, the reconciliation quantifies the gap; "
            "where only OAG is available this exclusion is stated on the face of the output, "
            "not in a footnote.")


OAG_STATEMENT = completeness_statement(
    "OAG published schedules (service type J), Sabre-weighted where pax basis",
    ["cargo", "general aviation", "military", "positioning", "ad-hoc charter"])


def reconcile(ap, date, towerlog_events, oag_events=None,
              oag_base_label="OAG monthly-file schedules, same calendar day"):
    """Side-by-side: towerlog actual day vs OAG schedule day, gap by category."""
    if oag_events is None:
        import ddfs_service as svc
        oag = svc.day_events(ap, date)
    else:
        oag = oag_events
    tl_hour = defaultdict(lambda: [0, 0])
    oag_hour = defaultdict(lambda: [0, 0])
    for e in towerlog_events:
        tl_hour[e["minute"] // 60][0 if e["ad"] == "A" else 1] += 1
    for e in oag:
        oag_hour[e["minute"] // 60][0 if e["ad"] == "A" else 1] += 1
    by_cls = defaultdict(int)
    for e in towerlog_events:
        by_cls[e["cls"]] += 1
    schedulable = by_cls["scheduled_pax"]
    gap = len(towerlog_events) - len(oag)
    return {
        "airport": ap, "date": str(date),
        "towerlog_movements": len(towerlog_events),
        "oag_movements": len(oag),
        "movement_gap": gap,
        "towerlog_by_class": dict(by_cls),
        "beyond_schedule": len(towerlog_events) - schedulable,
        "oag_vs_scheduled_pax": {"oag": len(oag), "towerlog_scheduled_pax": schedulable,
                                 "diff": schedulable - len(oag),
                                 "note": "residual = cancellations, ad-hoc retimes and schedule-vs-operated variance"},
        "hourly": {f"{h:02d}": {"towerlog": tl_hour[h], "oag": oag_hour[h]} for h in range(24)},
        "peak_hour": {"towerlog": max((sum(v) for v in tl_hour.values()), default=0),
                      "oag": max((sum(v) for v in oag_hour.values()), default=0)},
        "statement": OAG_STATEMENT,
        "oag_base": oag_base_label,
        "source": "MZLZ AODB towerlog day vs oag.duckdb schedules (" + oag_base_label + ")"}


def selftest():
    checks = []
    ev = load_fixture_day("zag_towerlog_sample_2025-05-26.tsv")
    checks.append(("sample rows parsed", len(ev), 14))
    by = defaultdict(int)
    for e in ev:
        by[e["cls"]] += 1
    checks.append(("scheduled pax", by["scheduled_pax"], 7))
    checks.append(("cargo", by["cargo"], 3))
    checks.append(("GA", by["ga"], 2))
    checks.append(("military (the C130)", by["military"], 1))
    checks.append(("positioning (ELP)", by["positioning"], 1))
    c130 = [e for e in ev if e["actype"] == "C130"][0]
    checks.append(("C130 classed military via Kind MC", c130["cls"], "military"))
    checks.append(("bizjet C25A maps ICAO A", [e for e in ev if e["actype"] == "C25A"][0]["icao"], "A"))
    checks.append(("minute parse (00:08 arrival)", min(e["minute"] for e in ev), 8))
    checks.append(("statement names the base", "OAG published schedules" in OAG_STATEMENT, True))
    checks.append(("statement names every exclusion",
                   all(x in OAG_STATEMENT for x in ("cargo", "general aviation", "military", "positioning")), True))
    # full-day pins (26 May 2025, extracted from the workbook 23 July 2026;
    # present once the extraction has run, checked whenever held)
    full_fx = os.path.join(FIX, "zag_towerlog_2025-05-26.tsv")
    if os.path.exists(full_fx):
        fev = load_fixture_day("zag_towerlog_2025-05-26.tsv")
        fby = defaultdict(int)
        for e in fev:
            fby[e["cls"]] += 1
        checks.append(("full day movements (pin)", len(fev), 153))
        checks.append(("full day scheduled pax (pin)", fby["scheduled_pax"], 126))
        checks.append(("full day beyond-schedule (pin)", len(fev) - fby["scheduled_pax"], 27))
        fk = {(e["ad"], e["minute"], e["registration"], e["fno"]) for e in fev}
        checks.append(("sample contained in full day",
                       sum(1 for e in ev if (e["ad"], e["minute"], e["registration"], e["fno"]) in fk), 14))
    fails = [c for c in checks if c[1] != c[2]]
    for c in checks:
        print(("ok  " if c[1] == c[2] else "FAIL"), c[0], c[1], "expected", c[2])
    print(f"towerlog selftest: {len(checks)} checks, " + ("all pass" if not fails else f"{len(fails)} FAILURES"))
    return not fails


def _local_xlsx():
    for pat in ("/sessions/*/mnt/C:--Avia/2025 Towerlog.xlsx",):
        hits = [h for h in glob.glob(pat) if os.access(h, os.R_OK)]
        if hits:
            return hits[0]
    p = "C:\\Avia\\2025 Towerlog.xlsx"
    return p if os.path.exists(p) else None


if __name__ == "__main__":
    if "--selftest" in sys.argv:
        sys.exit(0 if selftest() else 1)
    if "--extract" in sys.argv:
        i = sys.argv.index("--extract")
        extract_day(sys.argv[i + 1], dt.date.fromisoformat(sys.argv[i + 2]))
        sys.exit(0)
    if "--reconcile" in sys.argv:
        i = sys.argv.index("--reconcile")
        ap, d = sys.argv[i + 1], dt.date.fromisoformat(sys.argv[i + 2])
        fx = os.path.join(FIX, f"zag_towerlog_{d.isoformat()}.tsv")
        if os.path.exists(fx):
            ev = load_fixture_day(os.path.basename(fx))
        else:
            x = _local_xlsx()
            if not x:
                print("FLAG: no towerlog held (fixture or C:\\Avia\\2025 Towerlog.xlsx); cannot reconcile")
                sys.exit(1)
            ev = load_xlsx_day(x, d)
        import json
        from ddfs_service import _oag_day_or_wcweek
        oe, lab = _oag_day_or_wcweek(ap, d)
        r = reconcile(ap, d, ev, oag_events=oe, oag_base_label=lab)
        print(json.dumps({k: v for k, v in r.items() if k != "hourly"}, indent=1))
        print("hour  towerlog A/D   OAG A/D")
        for h in range(24):
            v = r["hourly"][f"{h:02d}"]
            print(f"{h:02d}    {v['towerlog'][0]:3d}/{v['towerlog'][1]:3d}      {v['oag'][0]:3d}/{v['oag'][1]:3d}")
